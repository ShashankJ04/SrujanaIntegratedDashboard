from __future__ import annotations

import io
from typing import Any, Dict, List, Optional, Set

from flask import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ColumnMeta, get_dashboard_rows_with_buffer

# Match static/js/table.js STATUS_COLS + row "ready" highlight
_STATUS_COLUMNS: Set[str] = {"production_pending", "balance_production_qty"}

_FILL_HEADER = PatternFill(
    start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"
)
_FILL_ROW_READY = PatternFill(
    start_color="ECFDF5", end_color="ECFDF5", fill_type="solid"
)
_FILL_CELL_GOOD = PatternFill(
    start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
)
_FILL_CELL_BAD = PatternFill(
    start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
)
_FONT_HEADER = Font(bold=True, size=12, color="0E2039")
_FONT_GOOD = Font(bold=True, color="0A7B4A")
_FONT_BAD = Font(bold=True, color="C0272D")
_FONT_DEFAULT = Font(color="0E2039")

# Thousands separator for numeric cells (Excel display)
_EXCEL_NUMERIC_FORMAT = "#,##0.00"


def _numeric_value_for_excel(val: Any) -> Any:
    """Return a float for any numeric dashboard value; bool and non-numeric stay unconverted."""
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _autosize_columns(ws, data: List[Dict[str, Any]], columns: List[ColumnMeta]) -> None:
    max_w = 55
    for idx, col in enumerate(columns, start=1):
        header_length = len(col.label)
        max_len = header_length
        for row in data:
            value = row.get(col.name)
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        adjusted_width = min(max_len + 2, max_w)
        ws.column_dimensions[get_column_letter(idx)].width = adjusted_width


def generate_excel_response(
    global_search: str,
    sort_by: str,
    sort_dir: str,
    row_filter: Optional[str] = None,
) -> Response:
    result = get_dashboard_rows_with_buffer(
        page=1,
        page_size=-1,
        global_search=global_search or None,
        sort_by=sort_by or None,
        sort_dir=sort_dir or None,
        row_filter=row_filter,
    )
    rows = result["rows"]
    columns_meta = [ColumnMeta(**c) for c in result["columns"]]

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    for col_idx, c in enumerate(columns_meta, start=1):
        cell = ws.cell(row=1, column=col_idx, value=c.label)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row_dict in enumerate(rows, start=2):
        prod_pending = float(row_dict.get("production_pending") or 0)
        balance_prod = float(row_dict.get("balance_production_qty") or 0)
        row_ready = prod_pending <= 0 and balance_prod <= 0

        for col_idx, col in enumerate(columns_meta, start=1):
            val = row_dict.get(col.name)
            cell = ws.cell(row=row_idx, column=col_idx)
            if col.is_numeric:
                num = _numeric_value_for_excel(val)
                if num is not None:
                    cell.value = num
                    cell.number_format = _EXCEL_NUMERIC_FORMAT
                else:
                    cell.value = val
            else:
                cell.value = val
            cell.font = _FONT_DEFAULT
            cell.alignment = Alignment(vertical="top")

            if row_ready:
                cell.fill = _FILL_ROW_READY

            if col.name in _STATUS_COLUMNS and val is not None:
                try:
                    num = float(val)
                except (TypeError, ValueError):
                    pass
                else:
                    if num <= 0:
                        cell.fill = _FILL_CELL_GOOD
                        cell.font = _FONT_GOOD
                    else:
                        cell.fill = _FILL_CELL_BAD
                        cell.font = _FONT_BAD

    ws.freeze_panes = "A2"
    _autosize_columns(ws, rows, columns_meta)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = "table_export.xlsx"
    response = Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
