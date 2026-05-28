"""Reports API Blueprint.

Port of dashboards/backend/src/routes/reports.ts.
"""

from __future__ import annotations

import re
from io import BytesIO
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file

_EXPORT_DT_FMT = "%d-%m-%Y, %H:%M:%S"
_EXCEL_NUMERIC_FORMAT = "#,##0.000000"
_EXCEL_HIGH_PRECISION_FORMAT = "#,##0." + ("0" * 10)


def _is_high_precision_column(col_name: str) -> bool:
    key = str(col_name or "").strip().lower()
    compact = key.replace(" ", "")
    if compact in ("rm_conval", "rmconval", "conval", "totalwt"):
        return True
    if "conval" in key or "con val" in key or "input rm" in key:
        return True
    return False

from .auth import api_login_required
from .rbac import require_access, require_plus_access
from . import reports_store
from .db import fetch_all

reports_bp = Blueprint("reports_bp", __name__, url_prefix="/api/reports")


@reports_bp.before_request
def _auth():
    result = api_login_required()
    if result is not None:
        return result


# ── Groups ──────────────────────────────────────────────────────────────

@reports_bp.route("/groups", methods=["GET"])
@require_access("rept")
def list_groups():
    return jsonify(reports_store.get_groups())


@reports_bp.route("/groups", methods=["POST"])
@require_plus_access("rept_plus")
def create_group():
    data = request.get_json(force=True)
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"message": "Name is required"}), 400
    try:
        group = reports_store.create_group(name)
        return jsonify(group), 201
    except ValueError as e:
        return jsonify({"message": str(e)}), 400


@reports_bp.route("/groups/<group_id>", methods=["DELETE"])
@require_plus_access("rept_plus")
def delete_group(group_id):
    try:
        reports_store.delete_group(group_id)
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"message": str(e)}), 404


# ── Reports ─────────────────────────────────────────────────────────────

@reports_bp.route("/reports", methods=["GET"])
@require_access("rept")
def list_reports():
    group_id = request.args.get("groupId")
    pinned_only = request.args.get("pinnedOnly", "").strip().lower() in ("1", "true", "yes")
    return jsonify(reports_store.get_reports(group_id, pinned_only=pinned_only))


@reports_bp.route("/reports/<report_id>", methods=["GET"])
@require_access("rept")
def get_report(report_id):
    report = reports_store.get_report_by_id(report_id)
    if not report:
        return jsonify({"message": "Report not found"}), 404
    return jsonify(report)


@reports_bp.route("/reports", methods=["POST"])
@require_plus_access("rept_plus")
def create_report():
    data = request.get_json(force=True)
    try:
        report = reports_store.create_report(
            group_id=data.get("groupId", ""),
            name=data.get("name", ""),
            query_template=data.get("queryTemplate", ""),
            drilldowns=data.get("drilldowns", []),
            pinned=bool(data.get("pinned", False)),
        )
        return jsonify(report), 201
    except ValueError as e:
        return jsonify({"message": str(e)}), 400


@reports_bp.route("/reports/<report_id>", methods=["PATCH"])
@require_plus_access("rept_plus")
def update_report(report_id):
    data = request.get_json(force=True)
    try:
        report = reports_store.update_report(
            report_id=report_id,
            name=data.get("name", ""),
            query_template=data.get("queryTemplate", ""),
            drilldowns=data.get("drilldowns", []),
            pinned=(
                bool(data.get("pinned"))
                if isinstance(data, dict) and "pinned" in data
                else None
            ),
        )
        return jsonify(report)
    except ValueError as e:
        return jsonify({"message": str(e)}), 400


@reports_bp.route("/reports/<report_id>", methods=["DELETE"])
@require_plus_access("rept_plus")
def delete_report(report_id):
    try:
        reports_store.delete_report(report_id)
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"message": str(e)}), 404


def _pascal_case_param_name(name: str) -> str:
    """Display param keys as PascalCase (e.g. start_date → StartDate, month → Month)."""
    s = str(name).strip()
    if not s:
        return s
    parts = [p for p in re.split(r"[_\s\-]+", s) if p]
    if not parts:
        return s
    return "".join(p[:1].upper() + p[1:].lower() for p in parts)


def _export_parameter_suffix(report: dict, variables: dict | None) -> str:
    """Comma-separated name: value pairs for the export title line."""
    variables = variables or {}
    names = list(report.get("variables") or [])
    seen: set[str] = set()
    parts: list[str] = []

    for name in names:
        if name in seen:
            continue
        seen.add(name)
        if name not in variables:
            continue
        val = variables[name]
        if val is None or str(val).strip() == "":
            continue
        parts.append(f"{_pascal_case_param_name(name)}: {val}")

    for name, val in sorted(variables.items()):
        if name in seen:
            continue
        if val is None or str(val).strip() == "":
            continue
        parts.append(f"{_pascal_case_param_name(name)}: {val}")

    if not parts:
        return ""
    return " — " + ", ".join(parts)


def _export_title_line(
    report_name: str,
    report: dict,
    variables: dict | None = None,
    exported_at: datetime | None = None,
) -> str:
    when = exported_at or datetime.now()
    base = f"{report_name} — Exported on {when.strftime(_EXPORT_DT_FMT)}"
    return base + _export_parameter_suffix(report, variables)


# ── Run report ──────────────────────────────────────────────────────────

@reports_bp.route("/reports/<report_id>/run", methods=["POST"])
@require_access("rept")
def run_report(report_id):
    report = reports_store.get_report_by_id(report_id)
    if not report:
        return jsonify({"message": "Report not found"}), 404

    data = request.get_json(force=True, silent=True) or {}
    variables = data.get("variables", {})

    try:
        sql, params = reports_store.compile_report_query(
            report["queryTemplate"], variables
        )
    except ValueError as e:
        return jsonify({"message": str(e)}), 400

    try:
        rows = fetch_all(sql, tuple(params))
    except Exception as e:
        return jsonify({"message": f"Query error: {str(e)}"}), 400

    columns = list(rows[0].keys()) if rows else []

    return jsonify({
        "reportId": report_id,
        "columns": columns,
        "rows": rows,
        "rowCount": len(rows),
        "executedAt": datetime.utcnow().isoformat(),
        "drilldowns": report.get("drilldowns", []),
    })


# ── Export report to Excel ──────────────────────────────────────────────

@reports_bp.route("/reports/<report_id>/export", methods=["POST"])
@require_access("rept")
def export_report(report_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    report = reports_store.get_report_by_id(report_id)
    if not report:
        return jsonify({"message": "Report not found"}), 404

    data = request.get_json(force=True, silent=True) or {}
    variables = data.get("variables", {})
    file_name = data.get("fileName", f"{report['name']}.xlsx")

    try:
        sql, params = reports_store.compile_report_query(
            report["queryTemplate"], variables
        )
    except ValueError as e:
        return jsonify({"message": str(e)}), 400

    try:
        rows = fetch_all(sql, tuple(params))
    except Exception as e:
        return jsonify({"message": f"Query error: {str(e)}"}), 400

    columns = list(rows[0].keys()) if rows else []
    col_count = max(len(columns), 1)

    wb = Workbook()
    ws = wb.active
    ws.title = report["name"][:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=13)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    next_row = 1
    title_cell = ws.cell(
        row=next_row, column=1,
        value=_export_title_line(report["name"], report, variables),
    )
    title_cell.font = title_font
    if col_count > 1:
        ws.merge_cells(
            start_row=next_row, start_column=1,
            end_row=next_row, end_column=col_count,
        )
    next_row += 2  # title row + blank row before table header

    header_row = next_row
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for ri, row in enumerate(rows, header_row + 1):
        for ci, col in enumerate(columns, 1):
            val = row.get(col)
            cell = ws.cell(row=ri, column=ci)
            if val is not None and not isinstance(val, bool):
                try:
                    num = float(val)
                    cell.value = num
                    cell.number_format = (
                        _EXCEL_HIGH_PRECISION_FORMAT
                        if _is_high_precision_column(col)
                        else _EXCEL_NUMERIC_FORMAT
                    )
                except (TypeError, ValueError):
                    cell.value = val
            else:
                cell.value = val
            cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=file_name,
    )
