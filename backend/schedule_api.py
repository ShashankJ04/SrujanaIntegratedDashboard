"""Schedule API Blueprint.

Port of dashboards/backend/src/routes/schedule.ts.
Corrected table names: schedule_master, schedule_details, scheduled_customer,
components (CO_), customer (CU_).
"""

from __future__ import annotations

import calendar
from datetime import datetime
from io import BytesIO

from flask import Blueprint, send_file

from .auth import api_login_required
from .db import fetch_all

schedule_bp = Blueprint("schedule_bp", __name__, url_prefix="/schedule")


@schedule_bp.before_request
def _auth():
    result = api_login_required()
    if result is not None:
        return result


@schedule_bp.route("/<date_param>", methods=["GET"])
def schedule_export(date_param):
    """Export schedule for a month as Excel using Customer Schedule pivot."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if len(date_param) != 6 or not date_param.isdigit():
        return {"message": "Invalid date format. Expected MMYYYY"}, 400

    mm = int(date_param[0:2])
    yyyy = int(date_param[2:6])

    if mm < 1 or mm > 12:
        return {"message": "Month must be between 01 and 12"}, 400

    days_in_month = calendar.monthrange(yyyy, mm)[1]

    rows = fetch_all(
        """
        SELECT
            cu.CU_Name      AS customer,
            c.CO_PARTNO     AS partno,
            DAY(sc.CS_DATE) AS day_num,
            SUM(sc.CS_QTY)  AS total_qty
        FROM schedule_master sm
            JOIN schedule_details sd   ON sm.SM_ID    = sd.SC_SMID
            JOIN scheduled_customer sc ON sd.SC_ID    = sc.CS_SCID
            JOIN components c          ON sd.SC_COMPID = c.CO_ID
            JOIN customer cu           ON c.CO_CUSTID  = cu.CU_Id
        WHERE sm.SM_MONTH = %s
            AND sm.SM_YEAR  = %s
            AND sc.CS_SCHEDULESTATE IN (1, 2)
        GROUP BY cu.CU_Name, c.CO_PARTNO, DAY(sc.CS_DATE)
        ORDER BY cu.CU_Name, c.CO_PARTNO, DAY(sc.CS_DATE)
        """,
        (mm, yyyy),
    )

    # Pivot: { "customer|partno" → { day → qty } }
    pivot = {}
    for r in rows:
        key = f"{r['customer']}|{r['partno']}"
        if key not in pivot:
            pivot[key] = {
                "customer": r["customer"] or "",
                "partno": r["partno"] or "",
                "days": {},
            }
        pivot[key]["days"][int(r["day_num"])] = int(r["total_qty"])

    # Grand totals
    grand_total_by_day = [0] * days_in_month
    grand_total = 0
    for entry in pivot.values():
        for d in range(1, days_in_month + 1):
            qty = entry["days"].get(d, 0)
            grand_total += qty
            grand_total_by_day[d - 1] += qty

    # Build Excel
    wb = Workbook()
    ws = wb.active

    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ]

    ws.title = f"Customer Schedule {mm:02d}-{yyyy}"

    # Title row
    now = datetime.now()
    ist_stamp = now.strftime("%d-%m-%Y %H:%M")
    title = f"Customer Schedule for {month_names[mm - 1]} {yyyy} as on {ist_stamp}"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days_in_month + 3)

    # Grand total row
    ws.cell(row=2, column=2, value="Grand Total").font = Font(bold=True)
    ws.cell(row=2, column=3, value=grand_total).font = Font(bold=True)
    for d in range(days_in_month):
        cell = ws.cell(row=2, column=d + 4, value=grand_total_by_day[d] or "")
        cell.font = Font(bold=True)

    # Header row
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_row = ["Customer", "Part No", "Total"] + [str(d) for d in range(1, days_in_month + 1)]
    for ci, h in enumerate(header_row, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10

    # Data rows
    row_idx = 4
    for entry in pivot.values():
        total = 0
        ws.cell(row=row_idx, column=1, value=entry["customer"])
        ws.cell(row=row_idx, column=2, value=entry["partno"])
        for d in range(1, days_in_month + 1):
            qty = entry["days"].get(d, 0)
            total += qty
            ws.cell(row=row_idx, column=d + 3, value=qty if qty else "")
        ws.cell(row=row_idx, column=3, value=total)
        row_idx += 1

    # ── Second sheet: PartNo Schedule ────────────────────────────────
    ws2 = wb.create_sheet(f"PartNo Schedule {mm:02d}-{yyyy}")

    # Aggregate by partno (no customer)
    part_pivot = {}
    for entry in pivot.values():
        partno = entry["partno"]
        if partno not in part_pivot:
            part_pivot[partno] = {}
        for d, qty in entry["days"].items():
            part_pivot[partno][d] = part_pivot[partno].get(d, 0) + qty

    ws2.cell(row=1, column=1, value=f"PartNo Schedule for {month_names[mm - 1]} {yyyy}").font = Font(bold=True, size=14)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days_in_month + 2)

    header2 = ["Part No", "Total"] + [str(d) for d in range(1, days_in_month + 1)]
    for ci, h in enumerate(header2, 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    row_idx = 3
    for partno in sorted(part_pivot.keys()):
        ws2.cell(row=row_idx, column=1, value=partno)
        total = 0
        for d in range(1, days_in_month + 1):
            qty = part_pivot[partno].get(d, 0)
            total += qty
            ws2.cell(row=row_idx, column=d + 2, value=qty if qty else "")
        ws2.cell(row=row_idx, column=2, value=total)
        row_idx += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"schedule_{mm:02d}{yyyy}.xlsx",
    )
