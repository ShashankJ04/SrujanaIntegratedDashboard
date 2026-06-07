"""Tool Breakdown API Blueprint."""

from __future__ import annotations

import calendar
from datetime import date, datetime
from functools import wraps
import math
from typing import Any, Dict, List, Optional

from io import BytesIO

from flask import Blueprint, jsonify, request, g, send_file

from .auth import api_login_required
from . import rbac
from .db import execute, fetch_all, fetch_one
from .pm_api import _norm_tool_no
from . import pm_store

tool_breakdowns_bp = Blueprint("tool_breakdowns_bp", __name__, url_prefix="/api/tool-breakdowns")

_BREAKDOWN_EXPORT_DT_FMT = "%d-%m-%Y, %H:%M:%S"
_BREAKDOWN_MONTH_ABBR = ("", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")


@tool_breakdowns_bp.before_request
def _auth():
    result = api_login_required()
    if result is not None:
        return result


def _current_perms() -> Dict[str, Any]:
    user = g.get("current_user") or {}
    return rbac.get_effective_permissions(
        user.get("userId", 0),
        user.get("login", ""),
        user.get("userId") == 43,
    )


def _require_perm(*, access: Optional[str] = None, plus: Optional[str] = None):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            perms = _current_perms()
            if access and access not in perms.get("access", []):
                return jsonify({"message": "Forbidden"}), 403
            if plus and plus not in perms.get("plusAccess", []):
                return jsonify({"message": "Forbidden"}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator


def _require_breakdown_list_access(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        perms = _current_perms()
        if (
            "preventive_maintenance" not in perms.get("access", [])
            and "edit_dpr" not in perms.get("plusAccess", [])
        ):
            return jsonify({"message": "Forbidden"}), 403
        return f(*args, **kwargs)
    return decorated


def _require_dpr_or_pm_access(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        perms = _current_perms()
        if (
            "preventive_maintenance" not in perms.get("access", [])
            and "edit_dpr" not in perms.get("plusAccess", [])
        ):
            return jsonify({"message": "Forbidden"}), 403
        return f(*args, **kwargs)
    return decorated


def _iso_dt(value: Any) -> Optional[str]:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _format_user_name(row: Dict[str, Any]) -> str:
    first = str(row.get("firstName") or "").strip()
    last = str(row.get("lastName") or "").strip()
    name = " ".join([p for p in (first, last) if p]).strip()
    login = str(row.get("login") or "").strip()
    return name or login


def _user_label(row: Dict[str, Any]) -> str:
    login = str(row.get("login") or "").strip()
    name = _format_user_name(row)
    if not name:
        return login
    if login and name.lower() != login.lower():
        return f"{name} ({login})"
    return name or login


def _fetch_active_user(user_id: Any) -> Optional[Dict[str, Any]]:
    try:
        uid = int(user_id)
    except (TypeError, ValueError):
        return None
    row = fetch_one(
        """
        SELECT
            US_ID AS id,
            US_Login AS login,
            COALESCE(US_FirstName, '') AS firstName,
            COALESCE(US_LastName, '') AS lastName
        FROM users
        WHERE US_CurrentYn = 'Y' AND US_ID = %s
        """,
        (uid,),
    )
    return row


def _operator_label(row: Dict[str, Any]) -> str:
    name = str(row.get("name") or "").strip()
    ecno = str(row.get("ecno") or "").strip()
    if name and ecno and name.lower() != ecno.lower():
        return f"{name} ({ecno})"
    return name or ecno


def _fetch_active_operator(operator_id: Any) -> Optional[Dict[str, Any]]:
    try:
        oid = int(operator_id)
    except (TypeError, ValueError):
        return None
    row = fetch_one(
        """
        SELECT
            OP_ID AS id,
            COALESCE(OP_ECNO, '') AS ecno,
            COALESCE(OP_NAME, '') AS name
        FROM operators
        WHERE OP_ACTIVEYN = 'Y' AND OP_ID = %s
        """,
        (oid,),
    )
    return row


def _strokes_from_produced_qty(part_no: Optional[str], produced_qty: Any) -> int:
    """Strokes for a produced qty on a part line (same formula as DPR / production_details)."""
    try:
        qty = float(produced_qty or 0)
    except (TypeError, ValueError):
        return 0
    if qty <= 0:
        return 0
    p = str(part_no or "").strip()
    if not p:
        return 0
    row = fetch_one(
        """
        SELECT ct.CT_NO_OF_CAVITY AS cavity
        FROM components_tool ct
        INNER JOIN components c ON ct.CT_COMPID = c.CO_ID
        WHERE ct.CT_ACTIVEYN = 'Y'
          AND c.CO_ACTIVEYN = 'Y'
          AND TRIM(c.CO_PARTNO) = %s
        ORDER BY ct.CT_ID DESC
        LIMIT 1
        """,
        (p,),
    )
    cavity = max(1, int(row.get("cavity") or 1)) if row else 1
    return int(round(qty / cavity))


def _as_of_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        try:
            return date.fromisoformat(value.strip()[:10])
        except ValueError:
            pass
    return date.today()


def _breakdown_completion_strokes(row: Dict[str, Any]) -> Dict[str, int]:
    """Current/next stroke at sign-off including DPR produced qty at and after breakdown."""
    tool_no = str(row.get("tool_no") or "").strip()
    tool_id = _resolve_tool_id_for_strokes(tool_no)
    if not tool_id:
        raise ValueError("Tool not found for stroke info")

    tl_row = fetch_one(
        "SELECT TL_preventive_maintenance_strokes FROM tool_life WHERE TL_tool_id = %s",
        (tool_id,),
    )
    if not tl_row:
        raise ValueError(f"Tool ID {tool_id} not found")
    pm_strokes = int(tl_row.get("TL_preventive_maintenance_strokes") or 0)

    as_of = _as_of_date(row.get("downtime_at"))
    erp_strokes = pm_store.get_tool_strokes(tool_id, as_of)

    part_no = str(row.get("part_no") or "").strip() or None
    dpr_strokes = 0
    q_at_breakdown = float(row.get("dpr_produced_qty") or 0)
    if part_no:
        dpr_strokes += _strokes_from_produced_qty(part_no, q_at_breakdown)
        dpr_row_id = row.get("dpr_row_id")
        if dpr_row_id is not None:
            try:
                dpr_id = int(dpr_row_id)
            except (TypeError, ValueError):
                dpr_id = None
            if dpr_id:
                dpr_row = fetch_one(
                    "SELECT produced_qty FROM dpr_daily_review WHERE id = %s",
                    (dpr_id,),
                )
                if dpr_row:
                    q_now = float(dpr_row.get("produced_qty") or 0)
                    delta = max(0.0, q_now - q_at_breakdown)
                    dpr_strokes += _strokes_from_produced_qty(part_no, delta)

    current = int(erp_strokes) + int(dpr_strokes)
    return {"currentStroke": current, "suggestedNextStroke": current + pm_strokes}


def _resolve_tool_id_for_strokes(tool_no: str) -> Optional[int]:
    norm = _norm_tool_no(tool_no).strip()
    if not norm:
        return None
    row = fetch_one(
        """
        SELECT TL_tool_id AS toolId
        FROM tool_life
        WHERE TL_tool_number = %s
        LIMIT 1
        """,
        (norm,),
    )
    if row and row.get("toolId") is not None:
        try:
            return int(row.get("toolId"))
        except (TypeError, ValueError):
            return None
    return None


def _breakdown_row_to_dict(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": row.get("id"),
        "toolNo": row.get("tool_no"),
        "partNo": row.get("part_no"),
        "partName": row.get("part_name"),
        "machineId": row.get("machine_id"),
        "machineName": row.get("machine_name"),
        "dprRowId": row.get("dpr_row_id"),
        "dprReviewDate": _iso_dt(row.get("dpr_review_date")),
        "dprProducedQty": row.get("dpr_produced_qty"),
        "issue": row.get("issue") or "",
        "priority": row.get("priority") or "Immediate",
        "operatorId": row.get("operator_user_id"),
        "operatorLogin": row.get("operator_login") or "",
        "operatorName": row.get("operator_name") or "",
        "downtimeAt": _iso_dt(row.get("downtime_at")),
        "rootCause": row.get("root_cause"),
        "rootCauseAt": _iso_dt(row.get("root_cause_at")),
        "analysis": row.get("analysis") or "",
        "actionTaken": row.get("action_taken"),
        "actionTakenAt": _iso_dt(row.get("action_taken_at")),
        "remarks": row.get("remarks") or "",
        "spareConsumed": row.get("spare_consumed") or "",
        "completedAt": _iso_dt(row.get("completed_at")),
        "completedById": row.get("completed_by_id"),
        "completedByLogin": row.get("completed_by_login"),
        "completedByName": row.get("completed_by_name"),
        "hoursSpentToConsume": row.get("hours_spent"),
        "currentStroke": row.get("current_stroke"),
        "nextStroke": row.get("next_stroke"),
        "createdBy": row.get("created_by"),
        "updatedBy": row.get("updated_by"),
        "createdAt": _iso_dt(row.get("created_at")),
        "updatedAt": _iso_dt(row.get("updated_at")),
        "status": "closed" if row.get("completed_at") else "active",
    }


@tool_breakdowns_bp.get("/operators")
@_require_perm(access="preventive_maintenance")
def list_breakdown_operators():
    rows = fetch_all(
        """
        SELECT
            US_ID AS id,
            US_Login AS login,
            COALESCE(US_FirstName, '') AS firstName,
            COALESCE(US_LastName, '') AS lastName
        FROM users
        WHERE US_CurrentYn = 'Y'
        ORDER BY US_Login
        """
    )
    result = []
    for r in rows:
        result.append(
            {
                "id": r.get("id"),
                "login": r.get("login"),
                "firstName": r.get("firstName"),
                "lastName": r.get("lastName"),
                "name": _format_user_name(r),
                "label": _user_label(r),
            }
        )
    return jsonify(result)


@tool_breakdowns_bp.get("/operators/dpr")
@_require_dpr_or_pm_access
def list_breakdown_operators_dpr():
    rows = fetch_all(
        """
        SELECT
            OP_ID AS id,
            COALESCE(OP_ECNO, '') AS ecno,
            COALESCE(OP_NAME, '') AS name
        FROM operators
        WHERE OP_ACTIVEYN = 'Y'
        ORDER BY OP_NAME, OP_ECNO
        """
    )
    result = []
    for r in rows:
        result.append(
            {
                "id": r.get("id"),
                "login": r.get("ecno"),
                "firstName": "",
                "lastName": "",
                "name": r.get("name"),
                "label": _operator_label(r),
            }
        )
    return jsonify(result)


@tool_breakdowns_bp.post("")
@_require_perm(plus="edit_dpr")
def create_breakdown():
    payload = request.get_json(silent=True) or {}
    tool_no_raw = payload.get("toolNo")
    tool_no = _norm_tool_no(tool_no_raw).strip()
    if not tool_no:
        return jsonify({"error": "toolNo is required"}), 400

    issue = str(payload.get("issue") or "").strip()
    if not issue:
        return jsonify({"error": "Issue/Problem is required"}), 400

    priority = str(payload.get("priority") or "Immediate").strip()
    if priority not in {"Immediate", "Next Day", "Delayed"}:
        return jsonify({"error": "Invalid priority"}), 400

    operator_id = payload.get("operatorId")
    operator_row = _fetch_active_operator(operator_id)
    if not operator_row:
        return jsonify({"error": "Invalid operator"}), 400

    open_row = fetch_one(
        "SELECT id FROM tool_breakdowns WHERE tool_no = %s AND completed_at IS NULL LIMIT 1",
        (tool_no,),
    )
    if open_row:
        return jsonify({"error": "An open breakdown already exists for this tool"}), 409

    part_no = str(payload.get("partNo") or "").strip() or None
    part_name = str(payload.get("partName") or "").strip() or None
    machine_id = str(payload.get("machineId") or "").strip() or None
    machine_name = str(payload.get("machineName") or "").strip() or None

    dpr_row_id: Optional[int] = None
    dpr_row_raw = payload.get("dprRowId")
    if dpr_row_raw not in (None, ""):
        try:
            dpr_row_id = int(dpr_row_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid dprRowId"}), 400

    dpr_review_date: Optional[date] = None
    dpr_review_raw = str(payload.get("dprReviewDate") or "").strip()
    if dpr_review_raw:
        try:
            dpr_review_date = date.fromisoformat(dpr_review_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid dprReviewDate"}), 400

    dpr_produced_qty: Optional[float] = None
    dpr_prod_raw = payload.get("dprProducedQty")
    if dpr_prod_raw not in (None, ""):
        try:
            dpr_produced_qty = float(dpr_prod_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid dprProducedQty"}), 400

    if dpr_row_id is not None:
        dpr_row = fetch_one(
            "SELECT id, produced_qty FROM dpr_daily_review WHERE id = %s",
            (dpr_row_id,),
        )
        if not dpr_row:
            return jsonify({"error": "DPR row not found"}), 404
        if dpr_row.get("produced_qty") is None:
            return jsonify({
                "error": "Enter Produced Qty on the DPR line before raising a breakdown (0 is allowed).",
            }), 400
        if dpr_produced_qty is None or not math.isfinite(dpr_produced_qty):
            return jsonify({
                "error": "Enter Produced Qty on the DPR line before raising a breakdown (0 is allowed).",
            }), 400
        dpr_produced_qty = float(dpr_row.get("produced_qty"))

    user_login = str(g.current_user.get("login") or "")
    execute(
        """
        INSERT INTO tool_breakdowns
            (tool_no, part_no, part_name, machine_id, machine_name,
             dpr_row_id, dpr_review_date, dpr_produced_qty, issue,
             priority, operator_user_id, operator_login, operator_name,
             downtime_at, created_by, updated_by)
        VALUES
            (%s, %s, %s, %s, %s,
             %s, %s, %s, %s,
             %s, %s, %s, %s,
             NOW(), %s, %s)
        """,
        (
            tool_no,
            part_no,
            part_name,
            machine_id,
            machine_name,
            dpr_row_id,
            dpr_review_date.isoformat() if dpr_review_date else None,
            dpr_produced_qty,
            issue,
            priority,
            operator_row.get("id"),
            operator_row.get("ecno"),
            operator_row.get("name"),
            user_login,
            user_login,
        ),
    )
    new_id_row = fetch_one("SELECT LAST_INSERT_ID() AS id")
    return jsonify({"id": new_id_row.get("id") if new_id_row else None})


_BREAKDOWN_SELECT_COLUMNS = """
            id, tool_no, part_no, part_name, machine_id, machine_name,
            dpr_row_id, dpr_review_date, dpr_produced_qty,
            issue, priority, operator_user_id, operator_login, operator_name, downtime_at,
            root_cause, root_cause_at, analysis, action_taken, action_taken_at, remarks, spare_consumed,
            completed_at, completed_by_id, completed_by_login, completed_by_name,
            hours_spent, current_stroke, next_stroke,
            created_by, updated_by, created_at, updated_at
"""


def _parse_month_year_filters() -> tuple[Optional[int], Optional[int], Optional[tuple]]:
    """Return (year, month, error_response) — filters on tool_breakdowns.created_at."""
    month_raw = request.args.get("month")
    year_raw = request.args.get("year")
    if month_raw in (None, "") and year_raw in (None, ""):
        return None, None, None
    if month_raw in (None, "") or year_raw in (None, ""):
        return None, None, (jsonify({"error": "month and year must be provided together"}), 400)
    try:
        month = int(month_raw)
        year = int(year_raw)
    except (TypeError, ValueError):
        return None, None, (jsonify({"error": "Invalid month or year"}), 400)
    if month < 1 or month > 12:
        return None, None, (jsonify({"error": "Invalid month"}), 400)
    if year < 2000 or year > 2100:
        return None, None, (jsonify({"error": "Invalid year"}), 400)
    return year, month, None


def _fetch_breakdown_rows(
    status: str,
    tool_no_raw: Optional[str] = None,
    limit_raw: Optional[str] = None,
) -> tuple[Optional[List[Dict[str, Any]]], Optional[tuple]]:
    where_parts: List[str] = []
    params: List[Any] = []
    if status == "closed":
        where_parts.append("completed_at IS NOT NULL")
    else:
        where_parts.append("completed_at IS NULL")

    filter_year, filter_month, period_err = _parse_month_year_filters()
    if period_err:
        return None, period_err
    if filter_year is not None and filter_month is not None:
        last_day = calendar.monthrange(filter_year, filter_month)[1]
        where_parts.append("DATE(created_at) >= %s AND DATE(created_at) <= %s")
        params.append(date(filter_year, filter_month, 1).isoformat())
        params.append(date(filter_year, filter_month, last_day).isoformat())

    if tool_no_raw:
        tool_no = _norm_tool_no(tool_no_raw).strip()
        if tool_no:
            where_parts.append("tool_no = %s")
            params.append(tool_no)

    limit_clause = ""
    if limit_raw not in (None, ""):
        try:
            lim = max(1, min(1000, int(limit_raw)))
            limit_clause = f" LIMIT {lim}"
        except (TypeError, ValueError):
            return None, (jsonify({"error": "Invalid limit"}), 400)

    where_sql = " AND ".join(where_parts) if where_parts else "1=1"
    sql = f"""
        SELECT
            {_BREAKDOWN_SELECT_COLUMNS}
        FROM tool_breakdowns
        WHERE {where_sql}
        ORDER BY downtime_at DESC, id DESC
        {limit_clause}
    """
    rows = fetch_all(sql, params if params else None)
    return rows, None


def _breakdown_export_title_line(
    status: str,
    month: Optional[int],
    year: Optional[int],
    exported_at: Optional[datetime] = None,
) -> str:
    status_label = "Closed" if status == "closed" else "Active"
    if month is not None and year is not None and 1 <= month <= 12:
        period = f"{_BREAKDOWN_MONTH_ABBR[month]}-{year}"
    else:
        period = "All Periods"
    when = exported_at or datetime.now()
    return (
        f"Tool Break Down {status_label} For {period} "
        f"Exported on {when.strftime(_BREAKDOWN_EXPORT_DT_FMT)}"
    )


def _breakdown_export_headers(*, closed: bool) -> List[str]:
    headers = [
        "Tool No",
        "Priority",
        "Part No",
        "Part Name",
        "Machine",
        "Downtime",
        "Produced Qty",
        "Issue / Problem",
        "Operator",
        "Root Cause",
        "Analysis",
        "Action Taken",
        "Remarks",
        "Spare Consumed",
        "Created At",
    ]
    if closed:
        headers.extend(["Hours Spent", "Completed By", "Completed At"])
    return headers


def _breakdown_export_row(rec: Dict[str, Any], *, closed: bool) -> List[Any]:
    machine = rec.get("machineName") or rec.get("machineId") or ""
    operator = rec.get("operatorName") or rec.get("operatorLogin") or ""
    completed_by = rec.get("completedByName") or rec.get("completedByLogin") or ""
    vals: List[Any] = [
        rec.get("toolNo") or "",
        rec.get("priority") or "Immediate",
        rec.get("partNo") or "",
        rec.get("partName") or "",
        machine,
        rec.get("downtimeAt") or "",
        rec.get("dprProducedQty") if rec.get("dprProducedQty") is not None else "",
        rec.get("issue") or "",
        operator,
        rec.get("rootCause") or "",
        rec.get("analysis") or "",
        rec.get("actionTaken") or "",
        rec.get("remarks") or "",
        rec.get("spareConsumed") or "",
        rec.get("createdAt") or "",
    ]
    if closed:
        vals.extend([
            rec.get("hoursSpentToConsume") if rec.get("hoursSpentToConsume") is not None else "",
            completed_by,
            rec.get("completedAt") or "",
        ])
    return vals


@tool_breakdowns_bp.get("")
@_require_breakdown_list_access
def list_breakdowns():
    status = str(request.args.get("status") or "active").strip().lower()
    rows, err = _fetch_breakdown_rows(
        status,
        tool_no_raw=request.args.get("toolNo"),
        limit_raw=request.args.get("limit"),
    )
    if err:
        return err
    return jsonify([_breakdown_row_to_dict(r) for r in rows or []])


@tool_breakdowns_bp.get("/export")
@_require_breakdown_list_access
def export_breakdowns():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    status = str(request.args.get("status") or "active").strip().lower()
    if status not in {"active", "closed"}:
        return jsonify({"error": "Invalid status"}), 400

    rows, err = _fetch_breakdown_rows(status)
    if err:
        return err

    closed = status == "closed"
    headers = _breakdown_export_headers(closed=closed)
    records = [_breakdown_row_to_dict(r) for r in rows or []]
    col_count = max(len(headers), 1)

    filter_year, filter_month, _ = _parse_month_year_filters()

    wb = Workbook()
    ws = wb.active
    ws.title = "Active Breakdowns" if not closed else "Closed Breakdowns"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=13)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    next_row = 1
    title_cell = ws.cell(
        row=next_row,
        column=1,
        value=_breakdown_export_title_line(status, filter_month, filter_year),
    )
    title_cell.font = title_font
    if col_count > 1:
        ws.merge_cells(
            start_row=next_row,
            start_column=1,
            end_row=next_row,
            end_column=col_count,
        )
    next_row += 2  # title row + blank row before column headers

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, rec in enumerate(records, next_row + 1):
        for col_idx, value in enumerate(_breakdown_export_row(rec, closed=closed), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    month_raw = request.args.get("month")
    year_raw = request.args.get("year")
    suffix = ""
    if month_raw and year_raw:
        try:
            suffix = f"_{int(year_raw)}_{int(month_raw):02d}"
        except (TypeError, ValueError):
            suffix = ""
    download_name = f"tool_breakdown{suffix}_{status}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name,
    )


@tool_breakdowns_bp.patch("/<int:breakdown_id>")
@_require_perm(plus="preventive_maintenance")
def update_breakdown(breakdown_id: int):
    payload = request.get_json(silent=True) or {}
    if (
        "rootCause" not in payload
        and "analysis" not in payload
        and "actionTaken" not in payload
        and "remarks" not in payload
        and "spareConsumed" not in payload
    ):
        return jsonify({"error": "No fields to update"}), 400

    existing = fetch_one(
        "SELECT id, completed_at FROM tool_breakdowns WHERE id = %s",
        (breakdown_id,),
    )
    if not existing:
        return jsonify({"error": "Breakdown not found"}), 404
    if existing.get("completed_at"):
        return jsonify({"error": "Completed breakdowns cannot be edited"}), 400

    set_parts: List[str] = []
    params: List[Any] = []

    if "rootCause" in payload:
        root_cause = str(payload.get("rootCause") or "").strip()
        if root_cause:
            set_parts.append("root_cause = %s")
            params.append(root_cause)
            set_parts.append("root_cause_at = NOW()")
        else:
            set_parts.append("root_cause = NULL")
            set_parts.append("root_cause_at = NULL")

    if "analysis" in payload:
        analysis = str(payload.get("analysis") or "").strip()
        if analysis:
            set_parts.append("analysis = %s")
            params.append(analysis)
        else:
            set_parts.append("analysis = NULL")

    if "actionTaken" in payload:
        action_taken = str(payload.get("actionTaken") or "").strip()
        if action_taken:
            set_parts.append("action_taken = %s")
            params.append(action_taken)
            set_parts.append("action_taken_at = NOW()")
        else:
            set_parts.append("action_taken = NULL")
            set_parts.append("action_taken_at = NULL")

    if "remarks" in payload:
        remarks = str(payload.get("remarks") or "").strip()
        if remarks:
            set_parts.append("remarks = %s")
            params.append(remarks)
        else:
            set_parts.append("remarks = NULL")

    if "spareConsumed" in payload:
        spare_consumed = str(payload.get("spareConsumed") or "").strip()
        if spare_consumed:
            set_parts.append("spare_consumed = %s")
            params.append(spare_consumed)
        else:
            set_parts.append("spare_consumed = NULL")

    set_parts.append("updated_by = %s")
    params.append(str(g.current_user.get("login") or ""))
    params.append(breakdown_id)

    execute(
        f"UPDATE tool_breakdowns SET {', '.join(set_parts)} WHERE id = %s",
        params,
    )
    row = fetch_one(
        """
        SELECT
            id, tool_no, part_no, part_name, machine_id, machine_name,
            dpr_row_id, dpr_review_date, dpr_produced_qty,
            issue, priority, operator_user_id, operator_login, operator_name, downtime_at,
            root_cause, root_cause_at, analysis, action_taken, action_taken_at, remarks, spare_consumed,
            completed_at, completed_by_id, completed_by_login, completed_by_name,
            hours_spent, current_stroke, next_stroke,
            created_by, updated_by, created_at, updated_at
        FROM tool_breakdowns
        WHERE id = %s
        """,
        (breakdown_id,),
    )
    return jsonify(_breakdown_row_to_dict(row) if row else {})


@tool_breakdowns_bp.post("/<int:breakdown_id>/complete")
@_require_perm(plus="preventive_maintenance")
def complete_breakdown(breakdown_id: int):
    payload = request.get_json(silent=True) or {}
    completed_by_id = payload.get("completedById")
    operator_row = _fetch_active_operator(completed_by_id)
    if not operator_row:
        return jsonify({"error": "Invalid completion user"}), 400

    row = fetch_one(
        """
        SELECT
            id, tool_no, part_no, dpr_row_id, dpr_produced_qty, downtime_at,
            root_cause, action_taken, completed_at
        FROM tool_breakdowns
        WHERE id = %s
        """,
        (breakdown_id,),
    )
    if not row:
        return jsonify({"error": "Breakdown not found"}), 404
    if row.get("completed_at"):
        return jsonify({"error": "Breakdown already completed"}), 400

    root_cause = str(row.get("root_cause") or "").strip()
    action_taken = str(row.get("action_taken") or "").strip()
    if not root_cause or not action_taken:
        return jsonify({"error": "Root cause and action taken are required"}), 400

    hours_raw = payload.get("hoursSpentToConsume")
    if hours_raw in (None, ""):
        return jsonify({"error": "Hours spent is required"}), 400
    try:
        hours_spent = float(hours_raw)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid hours spent"}), 400
    if hours_spent < 0:
        return jsonify({"error": "Hours spent cannot be negative"}), 400

    try:
        stroke_info = _breakdown_completion_strokes(row)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 404

    execute(
        """
        UPDATE tool_breakdowns
        SET
            completed_at = NOW(),
            completed_by_id = %s,
            completed_by_login = %s,
            completed_by_name = %s,
            hours_spent = %s,
            updated_by = %s,
            current_stroke = %s,
            next_stroke = %s
        WHERE id = %s
        """,
        (
            operator_row.get("id"),
            operator_row.get("ecno"),
            operator_row.get("name"),
            hours_spent,
            str(g.current_user.get("login") or ""),
            stroke_info.get("currentStroke"),
            stroke_info.get("suggestedNextStroke"),
            breakdown_id,
        ),
    )
    return jsonify({"ok": True})
