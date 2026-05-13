"""Preventive Maintenance API Blueprint.

Port of dashboards/backend/src/routes/pm.ts.
"""

from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path
from datetime import datetime
import threading
import time
from typing import Optional
from urllib.parse import unquote

from flask import Blueprint, jsonify, request, send_file, g, current_app

from .auth import api_login_required
from .rbac import require_access, require_plus_access, require_any_access
from . import pm_store

pm_bp = Blueprint("pm_bp", __name__, url_prefix="/api/pm")
_PM_STATUS_CACHE_LOCK = threading.Lock()
_PM_STATUS_CACHE = {}

# ── Aggregated: one row per tool (tool_life + latest PM + MAX strokes across parts).
# Core query matches git branch StableVersion1.7 backend/pm_api pm_status SQL.
# Used by GET /api/pm/status when per_component=0 (e.g. Life Report, Tools Today PM map).
_PM_STATUS_TOOL_AGG_SQL = """
        SELECT
            tl.TL_tool_id          AS toolId,
            tl.TL_tool_number      AS toolNo,
            tl.TL_life_span        AS toolLife,
            tl.TL_spm              AS spm,
            tl.TL_preventive_maintenance_strokes AS pmStrokes,
            pm.PM_next_stroke      AS nextStroke,
            pm.PM_date             AS lastMaintenanceDate,
            COALESCE(strokes.totalStrokes, 0) AS totalLifetimeStrokes,
            COALESCE(pm_count.cnt, 0)         AS maintenanceCount
        FROM tool_life tl
        INNER JOIN (
            SELECT pm1.*
            FROM preventive_maintenance pm1
            INNER JOIN (
                SELECT PM_tool_number, MAX(PM_id) AS maxId
                FROM preventive_maintenance
                GROUP BY PM_tool_number
            ) latest_pm ON latest_pm.PM_tool_number = pm1.PM_tool_number
                AND latest_pm.maxId = pm1.PM_id
        ) pm ON pm.PM_tool_number = tl.TL_tool_number
        LEFT JOIN (
            SELECT
                comp.toolNo,
                MAX(comp.componentStrokes) AS totalStrokes
            FROM (
                SELECT
                    ct.CT_TOOLNO AS toolNo,
                    ct.CT_COMPID,
                    SUM(pd.PD_PRODQTY / GREATEST(COALESCE(ct.CT_NO_OF_CAVITY, 1), 1)) AS componentStrokes
                FROM production_details pd
                INNER JOIN components_tool ct ON ct.CT_ID = pd.PD_TOOLID
                GROUP BY ct.CT_TOOLNO, ct.CT_COMPID
            ) comp
            GROUP BY comp.toolNo
        ) strokes ON strokes.toolNo = tl.TL_tool_number
        LEFT JOIN (
            SELECT PM_tool_number, COUNT(*) AS cnt
            FROM preventive_maintenance
            GROUP BY PM_tool_number
        ) pm_count ON pm_count.PM_tool_number = tl.TL_tool_number
        WHERE tl.TL_tool_id = (
            SELECT tl2.TL_tool_id
            FROM tool_life tl2
            WHERE tl2.TL_tool_number = tl.TL_tool_number
            ORDER BY tl2.TL_created_at DESC, tl2.TL_tool_id DESC
            LIMIT 1
        )
          AND EXISTS (
            SELECT 1
            FROM components_tool ct_active
            WHERE ct_active.CT_TOOLNO = tl.TL_tool_number
              AND ct_active.CT_ACTIVEYN = 'Y'
          )
        ORDER BY tl.TL_tool_number
        """

# Optional: one row per active components_tool; strokes = SUM for that PD_TOOLID (CT_ID).
# GET /api/pm/status?per_component=1 — used with aggregated status for PM hub slot expansion rule.
_PM_STATUS_COMPONENT_ROWS_SQL = """
        SELECT
            tl.TL_tool_id          AS toolId,
            ct.CT_TOOLNO           AS toolNo,
            COALESCE(c.CO_PARTNO, '') AS partNo,
            ct.CT_COMPID           AS compId,
            ct.CT_ID               AS componentToolId,
            tl.TL_life_span        AS toolLife,
            tl.TL_spm              AS spm,
            tl.TL_preventive_maintenance_strokes AS pmStrokes,
            pm.PM_next_stroke      AS nextStroke,
            pm.PM_date             AS lastMaintenanceDate,
            COALESCE(stroke_row.componentStrokes, 0) AS totalLifetimeStrokes,
            COALESCE(pm_count.cnt, 0)         AS maintenanceCount
        FROM tool_life tl
        INNER JOIN (
            SELECT pm1.*
            FROM preventive_maintenance pm1
            INNER JOIN (
                SELECT PM_tool_number, MAX(PM_id) AS maxId
                FROM preventive_maintenance
                GROUP BY PM_tool_number
            ) latest_pm ON latest_pm.PM_tool_number = pm1.PM_tool_number
                AND latest_pm.maxId = pm1.PM_id
        ) pm ON pm.PM_tool_number = tl.TL_tool_number
        INNER JOIN components_tool ct
            ON ct.CT_TOOLNO = tl.TL_tool_number
            AND ct.CT_ACTIVEYN = 'Y'
        INNER JOIN components c ON c.CO_ID = ct.CT_COMPID
        LEFT JOIN (
            SELECT
                pd.PD_TOOLID AS ctId,
                SUM(pd.PD_PRODQTY / GREATEST(COALESCE(ct2.CT_NO_OF_CAVITY, 1), 1)) AS componentStrokes
            FROM production_details pd
            INNER JOIN components_tool ct2 ON ct2.CT_ID = pd.PD_TOOLID
            GROUP BY pd.PD_TOOLID
        ) stroke_row ON stroke_row.ctId = ct.CT_ID
        LEFT JOIN (
            SELECT PM_tool_number, COUNT(*) AS cnt
            FROM preventive_maintenance
            GROUP BY PM_tool_number
        ) pm_count ON pm_count.PM_tool_number = tl.TL_tool_number
        WHERE tl.TL_tool_id = (
            SELECT tl2.TL_tool_id
            FROM tool_life tl2
            WHERE tl2.TL_tool_number = tl.TL_tool_number
            ORDER BY tl2.TL_created_at DESC, tl2.TL_tool_id DESC
            LIMIT 1
        )
        ORDER BY ct.CT_TOOLNO, c.CO_PARTNO, ct.CT_ID
        """

# One row per active components_tool slot (optional API / diagnostics only).
_PM_COMPONENT_TOOL_ROWS_SQL = """
        SELECT
            ct.CT_ID AS id,
            ct.CT_TOOLNO AS toolNo,
            COALESCE(c.CO_PARTNO, '') AS partNo
        FROM components_tool ct
        INNER JOIN components c ON c.CO_ID = ct.CT_COMPID
        WHERE ct.CT_ACTIVEYN = 'Y'
        ORDER BY ct.CT_TOOLNO, c.CO_PARTNO, ct.CT_ID
        """

# PM hub + export: one row per tool (matches dashboards/backend/src/routes/pm.ts export query).
_PM_GROUPED_TOOL_ROWS_SQL = """
        SELECT
            MIN(ct.CT_ID) AS id,
            ct.CT_TOOLNO AS toolNo,
            GROUP_CONCAT(DISTINCT c.CO_PARTNO ORDER BY c.CO_PARTNO SEPARATOR ', ') AS partNo
        FROM components_tool ct
        INNER JOIN components c ON c.CO_ID = ct.CT_COMPID
        WHERE ct.CT_ACTIVEYN = 'Y'
        GROUP BY ct.CT_TOOLNO
        ORDER BY ct.CT_TOOLNO
        """


def _norm_tool_no(value: object) -> str:
    import unicodedata

    s = unicodedata.normalize("NFKC", str(value or "").strip())
    return s


def _pm_entry_created_ts(e: dict) -> float:
    from datetime import datetime

    raw = e.get("createdAt") or ""
    if isinstance(raw, datetime):
        return raw.timestamp()
    try:
        return datetime.fromisoformat(str(raw).replace("Z", "+00:00")).timestamp()
    except Exception:
        return 0.0


def _canon_tool_key(tool_no: str, all_by_lower: dict[str, str]) -> str:
    n = _norm_tool_no(tool_no)
    return all_by_lower.get(n.lower(), n)


def _status_tool_key(value: object) -> str:
    """Hub / export join key: exact string like StableVersion1.7 JSON (no NFKC merge)."""
    if value is None:
        return ""
    return str(value)


def _pm_pct_for_status_row(st: dict) -> int:
    pm_strokes = int(st["pmStrokes"] or 0)
    next_stroke = int(st["nextStroke"] or 0)
    total_lifetime = int(st["totalLifetimeStrokes"] or 0)
    if pm_strokes <= 0:
        return 0
    cycle_start_stroke = next_stroke - pm_strokes
    completed_in_cycle = total_lifetime - cycle_start_stroke
    return max(0, round((completed_in_cycle / pm_strokes) * 100))


def _pm_merge_component_slots(slots: list[dict]) -> list[dict]:
    """Merge duplicate ``components_tool`` rows that share the same ``CT_COMPID`` (same component).

    Part labels can differ slightly in the DB; ``CT_COMPID`` is the stable join key to ``components``.
    Rows with missing ``compId`` fall back to one bucket per ``componentToolId`` (no merge).
    """
    if not slots:
        return []
    merged: dict[int, dict] = {}
    order: list[int] = []
    for r in slots:
        cid = int(r.get("compId") or 0)
        if cid <= 0:
            cid = -int(r.get("componentToolId") or 0)
        if cid not in merged:
            order.append(cid)
            m = dict(r)
            m["totalLifetimeStrokes"] = int(r.get("totalLifetimeStrokes") or 0)
            m["componentToolId"] = int(r["componentToolId"])
            merged[cid] = m
        else:
            m = merged[cid]
            m["totalLifetimeStrokes"] = int(m.get("totalLifetimeStrokes") or 0) + int(r.get("totalLifetimeStrokes") or 0)
            m["componentToolId"] = min(int(m["componentToolId"]), int(r["componentToolId"]))
    out = [merged[k] for k in order]
    out.sort(
        key=lambda x: (
            int(x.get("compId") or 0),
            _norm_tool_no(x.get("partNo")),
            int(x.get("componentToolId") or 0),
        )
    )
    return out


def _pm_merge_duplicate_part_labels(slots: list[dict]) -> list[dict]:
    """Second pass: merge rows whose part numbers normalize to the same display string.

    Some DBs have multiple ``components`` / ``CT_COMPID`` rows with the same ``CO_PARTNO``
    (or trivial spacing/case differences). ``CT_COMPID``-only merge leaves duplicates; this
    collapses them while summing strokes.
    """
    if len(slots) <= 1:
        return slots
    import re

    def part_key(r: dict) -> str:
        s = _norm_tool_no(r.get("partNo")).lower()
        return re.sub(r"\s+", " ", s).strip()

    merged: dict[str, dict] = {}
    order: list[str] = []
    for r in slots:
        pk = part_key(r)
        if pk not in merged:
            order.append(pk)
            m = dict(r)
            m["totalLifetimeStrokes"] = int(r.get("totalLifetimeStrokes") or 0)
            m["componentToolId"] = int(r["componentToolId"])
            merged[pk] = m
        else:
            m = merged[pk]
            m["totalLifetimeStrokes"] = int(m.get("totalLifetimeStrokes") or 0) + int(r.get("totalLifetimeStrokes") or 0)
            m["componentToolId"] = min(int(m["componentToolId"]), int(r["componentToolId"]))
    out = [merged[k] for k in order]
    out.sort(key=lambda x: (part_key(x), int(x.get("componentToolId") or 0)))
    return out


def _pm_dedupe_component_slot_rows(slots: list[dict]) -> list[dict]:
    """Merge duplicate CT rows (same ``CT_COMPID``), then merge same normalized part label."""
    return _pm_merge_duplicate_part_labels(_pm_merge_component_slots(slots))


def _is_path_under_root(root: str, candidate: str) -> bool:
    root = os.path.abspath(root)
    candidate = os.path.abspath(candidate)
    try:
        return os.path.commonpath([root, candidate]) == root
    except ValueError:
        return False


def _resolve_pm_attachment_file(full_dir: str, filename: str) -> Optional[str]:
    """Resolve a path or basename under ``full_dir``; search subdirs if direct path missing."""
    full_dir = os.path.abspath(full_dir)
    rel = unquote(filename).strip().replace("\\", "/").lstrip("/")
    if not rel or ".." in rel.split("/"):
        return None
    parts = [p for p in rel.split("/") if p and p != "."]
    if not parts:
        return None
    candidate = os.path.abspath(os.path.join(full_dir, *parts))
    if not _is_path_under_root(full_dir, candidate):
        return None
    if os.path.isfile(candidate):
        return candidate
    base = parts[-1]
    for p in Path(full_dir).rglob(base):
        if not p.is_file():
            continue
        rp = os.path.abspath(str(p))
        if _is_path_under_root(full_dir, rp) and p.name == base:
            return rp
    return None


def _invalidate_pm_status_cache() -> None:
    with _PM_STATUS_CACHE_LOCK:
        _PM_STATUS_CACHE.clear()


def _pm_attachments_root() -> str:
    """Absolute pm-attachments directory (uses app config — correct for PyInstaller exe)."""
    raw = current_app.config.get("PM_ATTACHMENTS_DIR")
    if raw:
        return os.path.abspath(os.path.expandvars(str(raw)))
    from .config import resolve_runtime_path

    return resolve_runtime_path(os.environ.get("PM_ATTACHMENTS_DIR", ""), "pm-attachments")


@pm_bp.before_request
def _auth():
    result = api_login_required()
    if result is not None:
        return result


# ── GET / — list all PM entries ─────────────────────────────────────────

@pm_bp.route("/", methods=["GET"])
@require_access("preventive_maintenance")
def list_entries():
    return jsonify(pm_store.get_entries())


# ── GET /status — PM status with percentage ─────────────────────────────

@pm_bp.route("/status", methods=["GET"])
@require_any_access(["preventive_maintenance", "life_report"])
def pm_status():
    from .db import fetch_all

    threshold = request.args.get("threshold", type=int, default=80)
    mode = request.args.get("mode", "above")
    per_component = str(request.args.get("per_component", "0")).lower() in ("1", "true", "yes")
    cache_seconds = int(current_app.config.get("PM_STATUS_CACHE_SECONDS", 20) or 20)
    cache_key = (10, int(threshold), str(mode), int(per_component))
    now = time.monotonic()
    if cache_seconds > 0:
        with _PM_STATUS_CACHE_LOCK:
            cached = _PM_STATUS_CACHE.get(cache_key)
            if cached and (now - cached["ts"]) < cache_seconds:
                return jsonify(cached["data"])

    sql = _PM_STATUS_COMPONENT_ROWS_SQL if per_component else _PM_STATUS_TOOL_AGG_SQL
    rows = fetch_all(sql)

    if per_component:
        _, all_by_lower = _all_active_tools_maps()
        tool_order: list[str] = []
        by_tool: dict[str, list[dict]] = {}
        for r in rows:
            tn = _canon_tool_key(_norm_tool_no(r.get("toolNo")), all_by_lower)
            if tn not in by_tool:
                tool_order.append(tn)
                by_tool[tn] = []
            by_tool[tn].append(r)
        rows = []
        for tn in tool_order:
            lst = by_tool[tn]
            lst.sort(
                key=lambda x: (
                    int(x.get("compId") or 0),
                    _norm_tool_no(x.get("partNo")),
                    int(x.get("componentToolId") or 0),
                )
            )
            rows.extend(_pm_dedupe_component_slot_rows(lst))

    results = []
    for r in rows:
        pm_pct = _pm_pct_for_status_row(r)
        if per_component:
            tool_no_out = _canon_tool_key(_norm_tool_no(r.get("toolNo")), all_by_lower)
        else:
            # StableVersion1.7: raw TL_tool_number (must match /api/tools/all CT_TOOLNO string-for-string for hub pills).
            tool_no_out = r.get("toolNo")

        entry = {
            "toolId": r["toolId"],
            "toolNo": tool_no_out,
            "toolLife": int(r["toolLife"]),
            "spm": int(r["spm"]),
            "pmStrokes": int(r["pmStrokes"] or 0),
            "pmCurrentStroke": int(r["nextStroke"] or 0) - int(r["pmStrokes"] or 0),
            "nextStroke": int(r["nextStroke"] or 0),
            "totalLifetimeStrokes": int(r["totalLifetimeStrokes"] or 0),
            "pmPercentage": pm_pct,
            "lastMaintenanceDate": str(r["lastMaintenanceDate"] or ""),
            "maintenanceCount": int(r["maintenanceCount"]),
        }
        if per_component:
            entry["partNo"] = _norm_tool_no(r.get("partNo"))
            entry["componentToolId"] = int(r["componentToolId"])
            entry["compId"] = int(r.get("compId") or 0)
        else:
            entry["partNo"] = str(r.get("partNo") or "").strip()

        if mode == "all":
            results.append(entry)
        elif pm_pct >= threshold:
            results.append(entry)

    if cache_seconds > 0:
        with _PM_STATUS_CACHE_LOCK:
            _PM_STATUS_CACHE[cache_key] = {"ts": now, "data": list(results)}
    return jsonify(results)


# ── GET /tool-strokes/<tool_id> ─────────────────────────────────────────

@pm_bp.route("/tool-strokes/<int:tool_id>", methods=["GET"])
@require_access("preventive_maintenance")
def tool_strokes(tool_id):
    total = pm_store.get_tool_strokes(tool_id)
    return jsonify({"totalStrokes": total})


# ── POST / — add entry ─────────────────────────────────────────────────

@pm_bp.route("/", methods=["POST"])
@require_plus_access("preventive_maintenance")
def add_entry():
    data = request.get_json(force=True)
    try:
        entry = pm_store.add_entry(
            tool_id=data["toolId"],
            tool_no=data["toolNo"],
            tool_life=data["toolLife"],
            spm=data["spm"],
            pm_strokes=data["pmStrokes"],
            next_stroke=data.get("nextStroke"),
        )
        _invalidate_pm_status_cache()
        return jsonify(entry), 201
    except ValueError as e:
        return jsonify({"message": str(e)}), 400


# ── PATCH /<tool_id> — update entry ─────────────────────────────────────

@pm_bp.route("/<int:tool_id>", methods=["PATCH"])
@require_plus_access("preventive_maintenance")
def update_entry(tool_id):
    data = request.get_json(force=True)
    try:
        entry = pm_store.update_entry(tool_id, data)
        _invalidate_pm_status_cache()
        return jsonify(entry)
    except ValueError as e:
        return jsonify({"message": str(e)}), 404


# ── GET /<tool_id>/stroke-info ──────────────────────────────────────────

@pm_bp.route("/<int:tool_id>/stroke-info", methods=["GET"])
@require_access("preventive_maintenance")
def stroke_info(tool_id):
    try:
        info = pm_store.get_stroke_info(tool_id)
        return jsonify(info)
    except ValueError as e:
        return jsonify({"message": str(e)}), 404


# ── POST /<tool_id>/confirm — confirm maintenance ──────────────────────

@pm_bp.route("/<int:tool_id>/confirm", methods=["POST"])
@require_plus_access("preventive_maintenance")
def confirm_maintenance(tool_id):
    next_stroke = request.form.get("nextStroke", type=int)
    if next_stroke is None:
        data = request.get_json(force=True, silent=True) or {}
        next_stroke = data.get("nextStroke")
    if next_stroke is None:
        return jsonify({"message": "nextStroke is required"}), 400

    attachment_name = None
    if "attachment" in request.files:
        f = request.files["attachment"]
        if f.filename:
            full_dir = _pm_attachments_root()
            os.makedirs(full_dir, exist_ok=True)
            safe_name = f"{tool_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{f.filename}"
            f.save(os.path.join(full_dir, safe_name))
            attachment_name = f"/api/pm/attachment/{safe_name}"

    try:
        entry = pm_store.confirm_maintenance(tool_id, next_stroke, attachment_name)
        _invalidate_pm_status_cache()
        return jsonify(entry)
    except ValueError as e:
        return jsonify({"message": str(e)}), 404


# ── DELETE /<tool_id> ───────────────────────────────────────────────────

@pm_bp.route("/<int:tool_id>", methods=["DELETE"])
@require_plus_access("preventive_maintenance")
def delete_entry(tool_id):
    try:
        pm_store.delete_entry(tool_id)
        _invalidate_pm_status_cache()
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"message": str(e)}), 404


# ── GET /attachment/<filename> ──────────────────────────────────────────

@pm_bp.route("/attachment/<path:filename>", methods=["GET"])
def serve_attachment(filename):
    full_dir = _pm_attachments_root()
    resolved = _resolve_pm_attachment_file(full_dir, filename)
    if not resolved:
        return jsonify({"message": "Attachment not found"}), 404
    return send_file(
        resolved,
        as_attachment=False,
        download_name=os.path.basename(resolved),
    )


def _all_active_tools_maps():
    """PM hub / export: one row per tool number (GROUP_CONCAT parts), like legacy pm.ts export."""
    from .db import fetch_all

    all_tool_rows = fetch_all(_PM_GROUPED_TOOL_ROWS_SQL)
    all_by_lower: dict[str, str] = {}
    for r in all_tool_rows:
        k = _norm_tool_no(r.get("toolNo"))
        if k:
            all_by_lower[k.lower()] = k
    return all_tool_rows, all_by_lower


def _load_pm_export_context():
    """Maps for PM hub + export: grouped tools + status joined by exact tool string (StableVersion1.7)."""
    from .db import fetch_all

    all_tool_rows, _ = _all_active_tools_maps()

    status_rows = fetch_all(_PM_STATUS_TOOL_AGG_SQL)
    status_by_exact: dict[str, dict] = {}
    for r in status_rows:
        k = _status_tool_key(r.get("toolNo"))
        if k:
            status_by_exact[k] = r

    entries = pm_store.get_entries()
    entry_by_exact: dict[str, dict] = {}
    for e in entries:
        k = _status_tool_key(e.get("toolNo"))
        if not k:
            continue
        ex = entry_by_exact.get(k)
        if ex is None or _pm_entry_created_ts(e) >= _pm_entry_created_ts(ex):
            entry_by_exact[k] = e

    return all_tool_rows, status_by_exact, entry_by_exact


def _build_pm_export_rows(mode: str, search: str) -> list[dict]:
    """Build Excel rows using the same rules as templates/hub_maintenance.html (no toolNos).

    We intentionally ignore the legacy ``toolNos`` query param: long GET URLs are truncated
    by proxies/browsers, and comma-separated tool numbers break when a tool no. contains ','.
    """
    search_l = (search or "").strip().lower()
    all_tool_rows, status_by_exact, entry_by_exact = _load_pm_export_context()

    processed: list[dict] = []
    for r in all_tool_rows:
        ct_key = _status_tool_key(r.get("toolNo"))
        part_display = str(r.get("partNo") or "")
        tool_no_l = ct_key.lower()
        part_no_l = part_display.lower()
        if search_l and search_l not in tool_no_l and search_l not in part_no_l:
            continue

        st = status_by_exact.get(ct_key)
        pm_pct = _pm_pct_for_status_row(st) if st else 0
        if mode == "safe" and pm_pct >= 80:
            continue
        if mode == "warning" and not (80 <= pm_pct < 100):
            continue
        if mode == "critical" and pm_pct < 100:
            continue

        ent = entry_by_exact.get(ct_key)

        latest = None
        if ent and ent.get("maintenanceHistory"):
            hist = ent["maintenanceHistory"]
            if hist:
                latest = hist[-1]

        total_strokes = int(st["totalLifetimeStrokes"] or 0) if st else 0
        if ent and not st:
            total_strokes = 0

        maint_count = int(st["maintenanceCount"] or 0) if st else (
            len(ent["maintenanceHistory"]) if ent and ent.get("maintenanceHistory") else 0
        )

        last_maint = ""
        if st and st.get("lastMaintenanceDate"):
            last_maint = str(st["lastMaintenanceDate"] or "")
        elif latest and latest.get("date"):
            last_maint = str(latest.get("date") or "")

        next_pm_val = int(latest["nextStroke"]) if latest and latest.get("nextStroke") is not None else None
        if next_pm_val is None and st:
            next_pm_val = int(st["nextStroke"] or 0)

        processed.append({
            "toolNo": ct_key,
            "partNo": part_display,
            "toolLife": int(ent["toolLife"]) if ent else None,
            "spm": int(ent["spm"]) if ent else None,
            "pmStrokes": int(ent["pmStrokes"]) if ent else None,
            "totalLifetimeStrokes": total_strokes,
            "nextStroke": next_pm_val,
            "pmPercentage": pm_pct,
            "maintenanceCount": maint_count,
            "lastMaintenanceDate": last_maint,
        })
    return processed


def pm_export_hub_parity_counts(search: str = "") -> dict[str, int]:
    """Counts matching hub Safe/Warning/Critical pills (one bucket per grouped tool, StableVersion1.7)."""
    search_l = (search or "").strip().lower()
    all_tool_rows, status_by_exact, _ = _load_pm_export_context()

    safe = warn = crit = 0
    for r in all_tool_rows:
        ct_key = _status_tool_key(r.get("toolNo"))
        part_display = str(r.get("partNo") or "")
        tnl = ct_key.lower()
        pnl = part_display.lower()
        if search_l and search_l not in tnl and search_l not in pnl:
            continue
        st = status_by_exact.get(ct_key)
        p = _pm_pct_for_status_row(st) if st else 0
        if p >= 100:
            crit += 1
        elif p >= 80:
            warn += 1
        else:
            safe += 1
    return {"safe": safe, "warning": warn, "critical": crit, "all": safe + warn + crit}


# ── GET /export — Excel export ──────────────────────────────────────────

@pm_bp.route("/export", methods=["GET"])
@require_access("preventive_maintenance")
def export_pm():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    mode = request.args.get("mode", "all")
    search = request.args.get("search", "").strip()

    processed = _build_pm_export_rows(mode, search)

    wb = Workbook()
    ws = wb.active
    ws.title = "Preventive Maintenance"

    headers = [
        "Sl No", "Tool No", "Part No", "Tool Life", "SPM", "PM Strokes",
        "Total Strokes", "Next PM Stroke", "PM %", "Maintenance Count",
        "Last Maintenance",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, entry in enumerate(processed, 2):
        vals = [
            row_idx - 1,
            entry["toolNo"],
            entry["partNo"],
            entry["toolLife"] if entry["toolLife"] is not None else "",
            entry["spm"] if entry["spm"] is not None else "",
            entry["pmStrokes"] if entry["pmStrokes"] is not None else "",
            entry["totalLifetimeStrokes"],
            entry["nextStroke"] if entry["nextStroke"] is not None else "",
            entry["pmPercentage"],
            entry["maintenanceCount"],
            entry["lastMaintenanceDate"],
        ]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="preventive_maintenance.xlsx",
    )
