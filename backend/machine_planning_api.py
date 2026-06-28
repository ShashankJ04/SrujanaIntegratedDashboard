"""Machine Planning API Blueprint.

REST endpoints for the monthly machine-wise production plan section.
"""

from __future__ import annotations

import os

from flask import Blueprint, current_app, g, jsonify, request, send_file

from .auth import api_login_required
from .rbac import require_access, require_plus_access
from . import machine_planning as mp

machine_planning_bp = Blueprint(
    "machine_planning_bp", __name__, url_prefix="/api/machine-planning"
)


@machine_planning_bp.before_request
def _auth():
    result = api_login_required()
    if result is not None:
        return result


# ── GET /machines ────────────────────────────────────────────────────────

@machine_planning_bp.route("/machines", methods=["GET"])
@require_access("rept")
def list_machines():
    machines = mp.get_machines()
    return jsonify(machines)


# ── GET /plan?machine_id=...&month=2026-04 ──────────────────────────────

@machine_planning_bp.route("/plan", methods=["GET"])
@require_access("rept")
def get_plan():
    machine_id = request.args.get("machine_id", type=int)
    month = request.args.get("month", "").strip()
    if not machine_id or not month:
        return jsonify({"message": "machine_id and month are required"}), 400
    try:
        data = mp.get_plan(machine_id, month)
    except Exception as exc:
        return jsonify({"message": str(exc)}), 500
    return jsonify(data)


# ── POST /plan  (add new row) ───────────────────────────────────────────

@machine_planning_bp.route("/plan", methods=["POST"])
@require_access("rept")
@require_plus_access("edit_dpr")
def add_row():
    body = request.get_json(silent=True) or {}
    machine_id = body.get("machine_id")
    month = body.get("month", "").strip()
    part_number = body.get("part_number", "").strip()
    if not machine_id or not month or not part_number:
        return jsonify({"message": "machine_id, month and part_number are required"}), 400

    user = g.get("current_user") or {}
    try:
        mp_id = mp.add_plan_row(
            machine_id=int(machine_id),
            month_year=month,
            part_number=part_number,
            additional_qty=int(body.get("additional_qty") or 0),
            priority=int(body.get("priority") or 0),
            remarks=str(body.get("remarks") or ""),
            created_by=user.get("userId"),
        )
    except Exception as exc:
        msg = str(exc)
        if "Duplicate" in msg:
            return jsonify({"message": "This part is already in the plan for this machine/month."}), 409
        return jsonify({"message": msg}), 500

    return jsonify({"mp_id": mp_id}), 201


# ── PATCH /plan/<mp_id>  (inline edit) ──────────────────────────────────

@machine_planning_bp.route("/plan/<int:mp_id>", methods=["PATCH"])
@require_access("rept")
@require_plus_access("edit_dpr")
def update_row(mp_id: int):
    body = request.get_json(silent=True) or {}
    fields = {}
    for key in ("additional_qty", "priority"):
        if key in body:
            fields[key] = int(body[key])
    if "remarks" in body:
        fields["remarks"] = str(body["remarks"])
    if not fields:
        return jsonify({"message": "No valid fields to update"}), 400
    mp.update_plan_row(mp_id, **fields)
    return jsonify({"ok": True})


# ── DELETE /plan/<mp_id> ────────────────────────────────────────────────

@machine_planning_bp.route("/plan/<int:mp_id>", methods=["DELETE"])
@require_access("rept")
@require_plus_access("edit_dpr")
def delete_row(mp_id: int):
    mp.delete_plan_row(mp_id)
    return jsonify({"ok": True})


# ── GET /part-search?q=... ──────────────────────────────────────────────

@machine_planning_bp.route("/part-search", methods=["GET"])
@require_access("rept")
def part_search():
    q = request.args.get("q", "").strip()
    results = mp.search_parts(q)
    return jsonify(results)


# ── GET /export?month=2026-06 ─────────────────────────────────────────

@machine_planning_bp.route("/export", methods=["GET"])
@require_access("rept")
def export_excel():
    """Export all machine plans for a month as a multi-sheet Excel file.

    Layout mirrors the original "APR MC WISE PLAN" Excel:
    Row 1: Title (B1:J1 merged) + Doc No (K1) + value (L1)
    Row 2: Header labels with merges (A2:C2, D2:E2, F, G, H, I2:I3 merged, K, L)
    Row 3: Header values with merges (A3:C3, D3:E3, F, G, H, J=total_days, L=date)
    Row 4: Data column headers (A..L) with wrap + white font on dark bg
    Row 5+: Data rows
    """
    from datetime import datetime
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    month = request.args.get("month", "").strip()
    if not month:
        return jsonify({"message": "month is required (YYYY-MM)"}), 400

    machine_ids = mp.get_machines_with_plans(month)
    if not machine_ids:
        return jsonify({"message": "No plans found for this month"}), 404

    ym = month.split("-")
    month_dt = datetime(int(ym[0]), int(ym[1]), 1)

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    # Match on-screen Machine Planning header card + table styling
    fill_title = PatternFill(start_color="0D3B66", end_color="0D3B66", fill_type="solid")
    fill_label = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    fill_value = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_hdr = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_data = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_data_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    title_font = Font(name="Calibri", bold=True, size=24, color="FFFFFF")
    doc_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    lbl_font = Font(name="Calibri", bold=True, size=14, color="44546A")
    val_font = Font(name="Calibri", bold=True, size=14, color="0F172A")
    val_accent_font = Font(name="Calibri", bold=True, size=14, color="16A34A")
    hdr_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    data_font = Font(name="Calibri", size=12, color="0F172A")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_al = Alignment(horizontal="right", vertical="center")

    def _style_row(row_idx: int, fill, font, *, accent_cols=None):
        accent_cols = accent_cols or set()
        for ci in range(1, 13):
            c = ws.cell(row=row_idx, column=ci)
            c.fill = fill
            c.font = val_accent_font if ci in accent_cols else font
            c.alignment = center
            c.border = thin

    col_widths = {
        "A": 15, "B": 38, "C": 29, "D": 17, "E": 19,
        "F": 24, "G": 13, "H": 17, "I": 17, "J": 17,
        "K": 20, "L": 27,
    }

    col_headers = [
        "SL",
        "PART NUMBER",
        "PART NAME",
        "PRODN. PLAN QTY",
        "PRODUCED QTY",
        "ADDITIONAL QTY(IF ANY WITH DTD)",
        "PRIORITY",
        "SPM",
        "MAXIMUM PARTS / DAY (7 hrs considered)",
        "TOTAL DAYS REQUIRED TO COMPLETE THE PRODUCTION",
        "RM CODE",
        "PPC/PROD \nSUPERVISOR\nREMARKS",
    ]

    logo_path = os.path.join(current_app.static_folder, "img", "se-logo.png")

    wb = Workbook()
    wb.remove(wb.active)

    for machine_id in machine_ids:
        plan = mp.get_plan(machine_id, month)
        mc = plan.get("machine") or {}
        rows = plan.get("rows") or []

        sheet_name = str(mc.get("label") or f"Machine {machine_id}")[:31]
        ws = wb.create_sheet(title=sheet_name)

        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w

        if os.path.isfile(logo_path):
            logo = XLImage(logo_path)
            target_h = 40
            if logo.height:
                logo.width = int(logo.width * target_h / logo.height)
            logo.height = target_h
            ws.add_image(logo, "A1")

        # ── Row 1: Title + Doc No ──
        ws.row_dimensions[1].height = 46
        ws.merge_cells("B1:J1")
        title = ws["B1"]
        title.value = "MONTHLY PRODUCTION PLAN  - POWER PRESS SHOP"
        title.font = title_font
        title.alignment = center
        ws["K1"].value = "Doc No:"
        ws["K1"].font = doc_font
        ws["K1"].alignment = center
        ws["L1"].value = "SE-PRD-F-21"
        ws["L1"].font = doc_font
        ws["L1"].alignment = center
        for ci in range(1, 13):
            c = ws.cell(row=1, column=ci)
            c.fill = fill_title
            c.border = thin
            c.alignment = center
        ws["B1"].font = title_font
        ws["K1"].font = doc_font
        ws["L1"].font = doc_font

        # ── Row 2: Header labels ──
        ws.row_dimensions[2].height = 30
        ws.merge_cells("A2:C2")
        ws["A2"].value = "MACHINE NO"
        ws.merge_cells("D2:E2")
        ws["D2"].value = "MACHINE NAME"
        ws["F2"].value = "TON"
        ws["G2"].value = "UNIT"
        ws["H2"].value = "MONTH/YEAR"
        ws["K2"].value = "Rev No. / Date"
        ws["L2"].value = f"REV.NO.00 Dt {datetime.now().strftime('%d-%m-%Y')}"
        _style_row(2, fill_label, lbl_font)

        # ── Row 3: Header values ──
        ws.row_dimensions[3].height = 36
        ws.merge_cells("A3:C3")
        ws["A3"].value = mc.get("label", "")
        ws.merge_cells("D3:E3")
        ws["D3"].value = mc.get("make", "")
        ws["F3"].value = mc.get("capacity", "")
        ws["G3"].value = 2
        ws["H3"].value = month_dt
        ws["H3"].number_format = "MMM YYYY"

        earliest = None
        for r in rows:
            ca = r.get("created_at")
            if ca and (not earliest or ca < earliest):
                earliest = ca

        ws["J3"].value = plan.get("total_days_required", 0)
        ws["J3"].number_format = "0.0"
        if earliest:
            try:
                ws["L3"].value = datetime.fromisoformat(earliest.replace("Z", "+00:00"))
                ws["L3"].number_format = "DD-MM-YYYY"
            except (ValueError, TypeError):
                ws["L3"].value = earliest[:10] if earliest else ""
        _style_row(3, fill_value, val_font, accent_cols={10})

        # ── Row 4: Column headers ──
        ws.row_dimensions[4].height = 98
        for ci, header_text in enumerate(col_headers, 1):
            c = ws.cell(row=4, column=ci, value=header_text)
            c.font = hdr_font
            c.fill = fill_hdr
            c.alignment = center
            c.border = thin

        # ── Data rows (row 5+) ──
        for ri, row in enumerate(rows, start=5):
            ws.row_dimensions[ri].height = 43
            row_fill = fill_data_alt if (ri - 5) % 2 == 1 else fill_data
            vals = [
                row.get("sl_no", ""),
                row.get("part_number", ""),
                row.get("part_name", ""),
                row.get("production_pending", 0),
                row.get("produced_qty", 0),
                row.get("additional_qty", 0),
                row.get("priority", 0),
                row.get("spm", 0),
                row.get("max_parts_per_day", 0),
                row.get("days_required", 0),
                row.get("rm_code", ""),
                row.get("remarks", ""),
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = data_font
                c.fill = row_fill
                c.alignment = center if ci in (1, 7) else (right_al if ci in (4,5,6,8,9,10) else left_al)
                c.border = thin
                if isinstance(val, (int, float)) and ci in (4, 5, 6, 8, 9):
                    c.number_format = "#,##0"
                elif ci == 10 and isinstance(val, (int, float)):
                    c.number_format = "0.0"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"machine_planning_{month}.xlsx",
    )
