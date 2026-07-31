"""Laser Welding report Excel export — matches hub Reports export styling."""

from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from flask import Response, send_file

from . import laser_welding as lw

_EXPORT_DT_FMT = "%d-%m-%Y, %H:%M:%S"
_EXCEL_NUMERIC_FORMAT = "#,##0"
_EXPORT_LIMIT = 10000

_REPORT_NAMES: Dict[str, str] = {
    "history": "Activity",
    "stock": "Stock",
    "qa": "QA",
    "scrap": "Scrap",
}

_VARIABLE_ORDER: Dict[str, Tuple[str, ...]] = {
    "history": ("from", "to", "step", "q"),
    "stock": ("q",),
    "qa": ("from", "to", "step", "q"),
    "scrap": ("from", "to", "step", "q"),
}


def _pascal_case_param_name(name: str) -> str:
    s = str(name).strip()
    if not s:
        return s
    parts = [p for p in re.split(r"[_\s\-]+", s) if p]
    if not parts:
        return s
    return "".join(p[:1].upper() + p[1:].lower() for p in parts)


def _export_parameter_suffix(
    variable_names: Tuple[str, ...],
    variables: Dict[str, Any],
) -> str:
    parts: List[str] = []
    seen: set[str] = set()
    for name in variable_names:
        if name in seen:
            continue
        seen.add(name)
        val = variables.get(name)
        if val is None or str(val).strip() == "":
            continue
        label = _pascal_case_param_name(name)
        if name == "q":
            label = "Search"
        parts.append(f"{label}: {val}")
    return (" — " + ", ".join(parts)) if parts else ""


def _export_title_line(
    report_name: str,
    variables: Dict[str, Any],
    variable_names: Tuple[str, ...],
    exported_at: Optional[datetime] = None,
) -> str:
    when = exported_at or datetime.now()
    base = f"{report_name} — Exported on {when.strftime(_EXPORT_DT_FMT)}"
    return base + _export_parameter_suffix(variable_names, variables)


def _format_export_date(iso: Any) -> str:
    parsed = lw._parse_date(iso)
    if not parsed:
        return str(iso or "").strip()
    try:
        dt = datetime.strptime(parsed, "%Y-%m-%d").date()
        return dt.strftime("%d-%m-%Y")
    except ValueError:
        return parsed


def _operator_label(row: Dict[str, Any]) -> str:
    name = str(row.get("operatorName") or "").strip()
    ecno = str(row.get("operatorEcno") or "").strip()
    if name and ecno:
        return f"{name} ({ecno})"
    return name or ecno or ""


def _time_taken_hours(minutes: Any) -> Optional[float]:
    """Convert minutes to decimal hours for Excel (e.g. 90 -> 1.5)."""
    if minutes is None:
        return None
    try:
        m = int(minutes)
    except (TypeError, ValueError):
        return None
    if m <= 0:
        return None
    return round(m / 60.0, 2)


def _normalize_variables(
    report_type: str,
    raw: Optional[Dict[str, Any]],
) -> Dict[str, str]:
    variables = raw or {}
    out: Dict[str, str] = {}
    if report_type != "stock":
        if variables.get("from"):
            out["from"] = _format_export_date(variables["from"])
        if variables.get("to"):
            out["to"] = _format_export_date(variables["to"])
        step = str(variables.get("step") or "").strip()
        if step:
            out["step"] = lw.HISTORY_STEP_LABELS.get(step, step)
    q = str(variables.get("q") or "").strip()
    if q:
        out["q"] = q
    return out


def _activity_export_rows(rows: List[Dict[str, Any]]) -> Tuple[List[str], List[Dict[str, Any]]]:
    columns = [
        "Date",
        "Step",
        "Type",
        "Part / BOM",
        "Lot",
        "Operator",
        "Machine",
        "Qty",
        "QA",
        "Scrap",
        "Rework",
        "Time",
        "OT",
    ]
    out: List[Dict[str, Any]] = []
    for r in rows:
        ot = str(r.get("otFlag") or "").upper()
        out.append(
            {
                "Date": _format_export_date(r.get("workDate")),
                "Step": r.get("workflowLabel") or "",
                "Type": r.get("rowClass") or r.get("rowType") or "",
                "Part / BOM": r.get("label") or r.get("partNo") or r.get("bomNo") or "",
                "Lot": r.get("packLotNo") or r.get("lotNo") or "",
                "Operator": _operator_label(r),
                "Machine": r.get("machineName") or "",
                "Qty": int(r.get("inspectedQty") or 0),
                "QA": int(r.get("qaQty") or 0),
                "Scrap": int(r.get("scrapQty") or 0),
                "Rework": int(r.get("reworkQty") or 0),
                "Time": _time_taken_hours(r.get("timeTakenMinutes")),
                "OT": "OT" if ot == "Y" else "",
            }
        )
    return columns, out


def _stock_export_rows(rows: List[Dict[str, Any]]) -> Tuple[List[str], List[Dict[str, Any]]]:
    columns = [
        "Type",
        "Part / BOM",
        "Name",
        "Inspection Pending",
        "FG",
        "QA",
        "Scrap",
        "Rework Pending",
        "Packed",
        "Total",
    ]
    out: List[Dict[str, Any]] = []
    for r in rows:
        out.append(
            {
                "Type": r.get("rowType") or "",
                "Part / BOM": r.get("label") or r.get("partNo") or r.get("bomNo") or "",
                "Name": r.get("partName") or "",
                "Inspection Pending": int(r.get("inspection_pending") or 0),
                "FG": int(r.get("fg") or 0),
                "QA": int(r.get("qa") or 0),
                "Scrap": int(r.get("scrap") or 0),
                "Rework Pending": int(r.get("rework_pending") or 0),
                "Packed": int(r.get("packed") or 0),
                "Total": int(r.get("totalQty") or 0),
            }
        )
    return columns, out


def _qa_export_rows(rows: List[Dict[str, Any]]) -> Tuple[List[str], List[Dict[str, Any]]]:
    columns = [
        "Date",
        "Type",
        "Part / BOM",
        "Lot",
        "Supplier",
        "Step",
        "Inspected",
        "QA",
        "Operator",
        "QA Remarks",
    ]
    out: List[Dict[str, Any]] = []
    for r in rows:
        out.append(
            {
                "Date": _format_export_date(r.get("workDate")),
                "Type": r.get("rowClass") or r.get("rowType") or "",
                "Part / BOM": r.get("label") or r.get("partNo") or r.get("bomNo") or "",
                "Lot": r.get("lotNo") or "",
                "Supplier": r.get("supplierName") or "",
                "Step": r.get("workflowLabel") or "",
                "Inspected": int(r.get("inspectedQty") or 0),
                "QA": int(r.get("qaQty") or 0),
                "Operator": _operator_label(r),
                "QA Remarks": r.get("qaRemark") or "",
            }
        )
    return columns, out


def _scrap_export_rows(rows: List[Dict[str, Any]]) -> Tuple[List[str], List[Dict[str, Any]]]:
    columns = [
        "Date",
        "Type",
        "Part / BOM",
        "Lot",
        "Supplier",
        "Step",
        "Inspected",
        "Scrap",
        "Scrap Remarks",
        "Operator",
        "Machine",
    ]
    out: List[Dict[str, Any]] = []
    for r in rows:
        out.append(
            {
                "Date": _format_export_date(r.get("workDate")),
                "Type": r.get("rowClass") or r.get("rowType") or "",
                "Part / BOM": r.get("label") or r.get("partNo") or r.get("bomNo") or "",
                "Lot": r.get("lotNo") or "",
                "Supplier": r.get("supplierName") or "",
                "Step": r.get("workflowLabel") or "",
                "Inspected": int(r.get("inspectedQty") or 0),
                "Scrap": int(r.get("scrapQty") or 0),
                "Scrap Remarks": r.get("scrapRemark") or "",
                "Operator": _operator_label(r),
                "Machine": r.get("machineName") or "",
            }
        )
    return columns, out


_HANDLER_TO_TYPE: Dict[str, str] = {
    "lw_activity": "history",
    "lw_stock": "stock",
    "lw_qa": "qa",
    "lw_scrap": "scrap",
}


def map_report_variables(variables: Optional[Dict[str, Any]]) -> Dict[str, str]:
    raw = variables or {}
    return {
        "from": str(raw.get("from_date") or raw.get("from") or "").strip(),
        "to": str(raw.get("to_date") or raw.get("to") or "").strip(),
        "q": str(raw.get("q") or "").strip(),
    }


def handler_to_report_type(handler: str) -> str:
    key = str(handler or "").strip()
    report_type = _HANDLER_TO_TYPE.get(key)
    if not report_type:
        raise ValueError(f"Unknown report handler: {handler}")
    return report_type


def run_builtin_report(handler: str, variables: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    report_type = handler_to_report_type(handler)
    mapped = map_report_variables(variables)
    if report_type != "stock":
        if not lw._parse_date(mapped.get("from")) or not lw._parse_date(mapped.get("to")):
            raise ValueError("from_date and to_date are required (YYYY-MM-DD)")
    columns, rows = _fetch_report_rows(report_type, mapped, apply_step_filter=False)
    return {
        "columns": columns,
        "rows": rows,
        "rowCount": len(rows),
    }


def export_builtin_report(
    handler: str,
    report_name: str,
    variables: Optional[Dict[str, Any]] = None,
    file_name: Optional[str] = None,
) -> Response:
    report_type = handler_to_report_type(handler)
    mapped = map_report_variables(variables)
    export_vars = _normalize_variables(report_type, mapped)
    var_order = _VARIABLE_ORDER[report_type]
    if report_type != "stock":
        raw = variables or {}
        if not lw._parse_date(raw.get("from_date") or raw.get("from")) or not lw._parse_date(
            raw.get("to_date") or raw.get("to")
        ):
            raise ValueError("from_date and to_date are required (YYYY-MM-DD)")
    columns, rows = _fetch_report_rows(report_type, mapped, apply_step_filter=False)
    output = _build_workbook(report_name, columns, rows, export_vars, var_order)
    download_name = str(file_name or f"{report_name}.xlsx").strip() or f"{report_name}.xlsx"
    if not download_name.lower().endswith(".xlsx"):
        download_name += ".xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name,
    )


def _fetch_report_rows(
    report_type: str,
    raw_variables: Dict[str, Any],
    *,
    apply_step_filter: bool = True,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    q = str(raw_variables.get("q") or "").strip()
    date_from = str(raw_variables.get("from") or "").strip()
    date_to = str(raw_variables.get("to") or "").strip()
    step = str(raw_variables.get("step") or "").strip() if apply_step_filter else ""
    if report_type == "history":
        data = lw.get_action_history(
            date_from=date_from,
            date_to=date_to,
            q=q,
            step=step,
            limit=_EXPORT_LIMIT,
        )
        return _activity_export_rows(data.get("rows") or [])
    if report_type == "stock":
        data = lw.get_stock_report(q=q)
        return _stock_export_rows(data.get("rows") or [])
    if report_type == "qa":
        data = lw.get_qa_history(
            date_from=date_from,
            date_to=date_to,
            q=q,
            step=step,
            limit=_EXPORT_LIMIT,
        )
        return _qa_export_rows(data.get("rows") or [])
    if report_type == "scrap":
        data = lw.get_scrap_history(
            date_from=date_from,
            date_to=date_to,
            q=q,
            step=step,
            limit=_EXPORT_LIMIT,
        )
        return _scrap_export_rows(data.get("rows") or [])
    raise ValueError("Invalid report type")


def _build_workbook(
    report_name: str,
    columns: List[str],
    rows: List[Dict[str, Any]],
    variables: Dict[str, str],
    variable_names: Tuple[str, ...],
) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side

    col_count = max(len(columns), 1)
    wb = Workbook()
    ws = wb.active
    ws.title = report_name[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=13)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    next_row = 1
    title_cell = ws.cell(
        row=next_row,
        column=1,
        value=_export_title_line(report_name, variables, variable_names),
    )
    title_cell.font = title_font
    if col_count > 1:
        ws.merge_cells(
            start_row=next_row,
            start_column=1,
            end_row=next_row,
            end_column=col_count,
        )
    next_row += 2

    header_row = next_row
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for ri, row in enumerate(rows, header_row + 1):
        for ci, col in enumerate(columns, 1):
            val = row.get(col)
            cell = ws.cell(row=ri, column=ci)
            if val is not None and not isinstance(val, bool):
                try:
                    num = float(val)
                    if num == int(num):
                        cell.value = int(num)
                        cell.number_format = _EXCEL_NUMERIC_FORMAT
                    else:
                        cell.value = num
                        cell.number_format = "#,##0.####"
                except (TypeError, ValueError):
                    cell.value = val
            else:
                cell.value = val
            cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_lw_report_export(
    report_type: str,
    variables: Optional[Dict[str, Any]] = None,
    file_name: Optional[str] = None,
) -> Response:
    report_key = str(report_type or "").strip().lower()
    report_name = _REPORT_NAMES.get(report_key)
    if not report_name:
        raise ValueError("Invalid report type")

    export_vars = _normalize_variables(report_key, variables)
    var_order = _VARIABLE_ORDER[report_key]

    if report_key != "stock":
        raw = variables or {}
        if not lw._parse_date(raw.get("from")) or not lw._parse_date(raw.get("to")):
            raise ValueError("from and to dates are required (YYYY-MM-DD)")

    columns, rows = _fetch_report_rows(report_key, variables or {})
    output = _build_workbook(report_name, columns, rows, export_vars, var_order)
    download_name = str(file_name or f"{report_name}.xlsx").strip() or f"{report_name}.xlsx"
    if not download_name.lower().endswith(".xlsx"):
        download_name += ".xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name,
    )
